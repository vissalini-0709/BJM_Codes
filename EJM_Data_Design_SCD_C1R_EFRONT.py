#This code wil enhance, add column, red and bold rows with unmatch received and processed records.
#This code only for SCD,C1R and EFRONT


import pandas as pd
import os
from tkinter import Tk, filedialog
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

def get_interface_value(process_name, stream_name):
    p = str(process_name or '').lower()
    s = str(stream_name or '').lower()

    key_str = 'eagle_ml-2-0_default_in_xml_warehouse_preproc'

    # New mappings added here
    if 'inbnd_factset_r_smf_fi' in p:
        return 'C1R FI'
    if 'inbnd_factset_r_smf_eq' in p:
        return 'C1R EQ'
    if 'inbnd_efront_smf' in p:
        return 'EFRONT SMF'
    if 'inbnd_efront_txn' in p:
        return 'EFRONT TXN'
    if 'inbnd_efront_holdings' in p:
        return 'EFRONT HOLDINGS'

    # Existing mappings
    if 'inbnd_scd_smf' in p:
        return 'SMF'
    if 'inbnd_scd_fxrate' in p:
        return 'FX RATE'
    if 'inbnd_scd_secrating' in p:
        return 'SECURITY RATING'
    if 'inbnd_scd_capinfo' in p:
        return 'CAPITAL INFO'
    if 'inbnd_scd_holdings' in p:
        return 'Holdings'
    if 'inbnd_scd_glbal' in p:
        return 'GL Balances'
    if 'inbnd_scd_nav' in p:
        return 'NAV'
    if 'inbnd_scd_saataa' in p:
        return 'SAA TAA'

    # txn special case: count occurrences of key_str in stream_name
    if 'inbnd_scd_txn' in p:
        count = s.count(key_str)
        if count == 1:
            return 'Transaction PL'
        elif count >= 2:
            return 'Transaction Cost'
        else:
            return ''  # stream name does not contain key_str

    return ''

def parse_datetime(date_str, time_str):
    try:
        dt_str = f"{date_str} {time_str}".strip()
        for fmt in ("%d/%m/%y %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(dt_str, fmt)
            except:
                continue
        return None
    except:
        return None

def highlight_mismatch(received, processed):
    try:
        return int(received) != int(processed)
    except:
        return False

def parse_first_successful_total(stream_name):
    """
    Extract only the first (successful:x total:y) pair from the stream_name string.
    Returns (successful, total) as integers or (0,0) if not found.
    """
    if not isinstance(stream_name, str):
        return 0, 0
    match = re.search(r'successful:\s*(\d+)\s+total:\s*(\d+)', stream_name, re.IGNORECASE)
    if match:
        return int(match.group(1)), int(match.group(2))
    return 0, 0

def process_file(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    df.columns = df.columns.str.strip()

    # Compute Interface column per row using existing logic
    df['Interface'] = df.apply(lambda r: get_interface_value(r.get('Process Name', ''), r.get('Stream Name', '')), axis=1)

    # For rows with Interface == 'Transaction Cost', parse first (successful:x total:y) from Stream Name
    txn_cost_mask = df['Interface'] == 'Transaction Cost'
    for idx in df[txn_cost_mask].index:
        stream_val = df.at[idx, 'Stream Name']
        x, y = parse_first_successful_total(stream_val)
        # Overwrite Received Record and Processed Record columns
        df.at[idx, 'Received Record'] = y
        df.at[idx, 'Processed Record'] = x

    # Parse Date and Start Time for sorting
    df['ParsedDateTime'] = df.apply(lambda r: parse_datetime(str(r.get('Date', '') or ''), str(r.get('Start Time', '') or '')), axis=1)

    # Sort by Interface then by ParsedDateTime ascending (oldest first)
    df.sort_values(by=['Interface', 'ParsedDateTime'], inplace=True)

    # Move Interface column to first position
    cols = list(df.columns)
    if 'Interface' in cols:
        cols.insert(0, cols.pop(cols.index('Interface')))
    df = df[cols]

    # Prepare output file path
    output_folder = os.path.dirname(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_folder, f"Modified_Data_{timestamp}.xlsx")

    # Save dataframe to Excel (drop ParsedDateTime before saving)
    df.drop(columns=['ParsedDateTime'], inplace=True)
    df.to_excel(output_file, index=False)

    # Open with openpyxl to apply red bold font where Received Record != Processed Record
    wb = load_workbook(output_file)
    ws = wb.active

    # Find column indexes (1-based)
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        rec_col = header.index('Received Record') + 1
        proc_col = header.index('Processed Record') + 1
    except ValueError:
        rec_col = None
        proc_col = None

    red_bold_font = Font(color="FF0000", bold=True)

    if rec_col and proc_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            val_rec = row[rec_col - 1].value
            val_proc = row[proc_col - 1].value
            if highlight_mismatch(val_rec, val_proc):
                for cell in row:
                    cell.font = red_bold_font

    wb.save(output_file)
    print(f"Processed file saved to: {output_file}")

def main():
    Tk().withdraw()
    files = filedialog.askopenfilenames(title="Select Excel file(s)", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not files:
        print("No files selected.")
        return
    for f in files:
        print(f"Processing file: {f}")
        process_file(f)

if __name__ == "__main__":
    main()
